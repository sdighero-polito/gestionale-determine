"""
Database v7 — Gestione Contratti Demanio Marittimo, Comune di Sanremo
Schema basato su ER v7 con tre layer (amministrativo, contabile, tecnico),
macchina a stati, allegati polimorfici, log attività.
"""

import sqlite3
import os
import json
from datetime import datetime, date, timedelta
from contextlib import contextmanager

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "contratti.db")

# ═══════════════════════════════════════════════════════════════
# COSTANTI
# ═══════════════════════════════════════════════════════════════

TIPI_PROCEDURA = [
    "Affidamento diretto (art. 50 c.1 lett. a)",
    "Affidamento diretto (art. 50 c.1 lett. b)",
    "Procedura negoziata senza bando (art. 50 c.1 lett. c)",
    "Procedura negoziata senza bando (art. 50 c.1 lett. d)",
    "Procedura negoziata senza bando (art. 76)",
    "Procedura aperta (art. 71)",
    "Procedura ristretta (art. 72)",
    "Dialogo competitivo (art. 77)",
    "Partenariato per l'innovazione (art. 78)",
    "Procedura competitiva con negoziazione (art. 73)",
    "Concessione (art. 176 e ss.)",
    "Somma urgenza (art. 140)",
    "Altro",
]

TIPI_PRESTAZIONE = [
    "Lavori",
    "Servizi",
    "Forniture",
    "Servizi tecnici (ingegneria/architettura)",
    "Concessione demaniale",
    "Concessione SLA (art. 45-bis Cod. Nav.)",
]

TIPI_DETERMINA = [
    "Impegno/Accertamento",
    "Liquidazione",
    "Determina Dirigenziale"
]

RIF_NORMATIVI = [
    "Art. 50, comma 1, lett. a) D.Lgs. 36/2023 (Lavori < 150k)",
    "Art. 50, comma 1, lett. b) D.Lgs. 36/2023 (Servizi/Forniture < 140k)",
    "Art. 71 D.Lgs. 36/2023 (Procedura aperta)",
    "Art. 76 D.Lgs. 36/2023 (Procedura negoziata senza bando)",
    "D.Lgs. 36/2023",
    "Legge 241/1990",
]

STATI_ITER_DETERMINA = ["Bozza", "In firma", "Firmata", "Pubblicata"]

STATI_AFFIDAMENTO = [
    "Bozza",
    "In istruttoria",
    "In gara",
    "In affidamento diretto",
    "Aggiudicato non efficace",
    "Verifica requisiti",
    "Affidato",
    "In esecuzione",
    "Sospeso",
    "Concluso",
    "Liquidazione parziale",
    "Liquidato totale",
]

# Transizioni valide: stato_corrente -> [stati_possibili]
TRANSIZIONI_STATO = {
    "Bozza": ["In istruttoria"],
    "In istruttoria": ["In gara", "In affidamento diretto"],
    "In gara": ["Aggiudicato non efficace", "In istruttoria"],
    "In affidamento diretto": ["Aggiudicato non efficace", "In istruttoria"],
    "Aggiudicato non efficace": ["Verifica requisiti"],
    "Verifica requisiti": ["Affidato", "In istruttoria"],
    "Affidato": ["In esecuzione"],
    "In esecuzione": ["Sospeso", "Concluso"],
    "Sospeso": ["In esecuzione"],
    "Concluso": ["Liquidazione parziale", "Liquidato totale"],
    "Liquidazione parziale": ["Liquidazione parziale", "Liquidato totale"],
    "Liquidato totale": [],
}

TIPI_FASE_CONTABILE = ["Prenotazione", "Impegno", "Liquidazione", "Economia"]

TIPI_VERBALE = [
    "Consegna dei luoghi",
    "Sospensione lavori",
    "Ripresa lavori",
    "SAL",
    "Fine lavori",
    "CRE",
]

FASI_VERBALE = {"Consegna dei luoghi": "consegna", "Sospensione lavori": "esecuzione",
    "Ripresa lavori": "esecuzione", "SAL": "esecuzione", "Fine lavori": "chiusura", "CRE": "chiusura"}

TIPI_VERIFICA = [
    "DURC",
    "Requisiti generali (art. 94-98 D.Lgs. 36/2023)",
    "Certificato antimafia",
    "Casellario ANAC",
    "Regolarità fiscale",
    "Regolarità contributiva",
    "Altro",
]

CANALI_PUBBLICAZIONE = [
    "Albo Pretorio",
    "Maggioli (e-procurement)",
    "ANAC",
    "GUCE",
    "BURL",
    "Sito istituzionale",
]

SOGLIE_COMUNITARIE = {
    "Lavori": 5_382_000,
    "Servizi": 221_000,
    "Forniture": 221_000,
    "Servizi tecnici (ingegneria/architettura)": 221_000,
    "Concessione demaniale": 5_382_000,
    "Concessione SLA (art. 45-bis Cod. Nav.)": 5_382_000,
}

SOGLIE_AFFIDAMENTO_DIRETTO = {
    "Lavori": 150_000,
    "Servizi": 140_000,
    "Forniture": 140_000,
    "Servizi tecnici (ingegneria/architettura)": 140_000,
}

TIPI_GARA = [
    "Procedura aperta",
    "Procedura ristretta",
    "Procedura negoziata",
    "Dialogo competitivo",
    "Concessione demaniale marittima",
    "Concessione SLA",
]

REPORT_PREIMPOSTATI = {
    "Affidamenti attivi": "SELECT a.id, a.oggetto, a.tipo_procedura, a.stato, o.ragione_sociale, a.importo_affidato FROM affidamenti a LEFT JOIN operatori_economici o ON a.id_operatore=o.id WHERE a.stato NOT IN ('Liquidato totale','Bozza') ORDER BY a.id DESC",
    "Fatture da liquidare": "SELECT f.id, f.numero_fattura, f.data_fattura, o.ragione_sociale, f.importo_totale, f.stato FROM fatture f LEFT JOIN operatori_economici o ON f.id_operatore=o.id WHERE f.stato_pagamento='Da liquidare' ORDER BY f.data_fattura",
    "Determine per tipo": "SELECT tipo_determina, COUNT(*) as n, SUM(importo) as tot FROM determine GROUP BY tipo_determina ORDER BY n DESC",
    "Quadro economico per affidamento": "SELECT a.id, a.oggetto, SUM(CASE WHEN q.sezione='A' THEN q.importo ELSE 0 END) as sez_a, SUM(CASE WHEN q.sezione='B' THEN q.importo ELSE 0 END) as sez_b, SUM(q.importo) as totale FROM affidamenti a JOIN quadro_economico q ON q.id_affidamento=a.id GROUP BY a.id",
    "Situazione contabile": "SELECT a.id, a.oggetto, a.importo_affidato, COALESCE(SUM(CASE WHEN m.tipo_fase='Liquidazione' THEN m.importo END),0) as liquidato FROM affidamenti a LEFT JOIN movimenti_contabili m ON m.id_affidamento=a.id WHERE a.stato NOT IN ('Bozza') GROUP BY a.id",
    "Verifiche in scadenza": "SELECT c.tipo_verifica, c.data_scadenza, a.oggetto, c.esito FROM checklist_verifiche c JOIN affidamenti a ON c.id_affidamento=a.id WHERE c.data_scadenza IS NOT NULL AND c.data_scadenza <= date('now','+30 days') ORDER BY c.data_scadenza",
    "Gare in corso": "SELECT g.id, g.tipo_gara, g.oggetto, g.fase_corrente, g.importo_base_asta, g.stato FROM gare g WHERE g.stato NOT IN ('Completata','Annullata') ORDER BY g.id DESC",
}


# ═══════════════════════════════════════════════════════════════
# CONNESSIONE
# ═══════════════════════════════════════════════════════════════

def get_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


@contextmanager
def get_db():
    conn = get_connection()
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


# ═══════════════════════════════════════════════════════════════
# CREAZIONE SCHEMA
# ═══════════════════════════════════════════════════════════════

def init_db():
    conn = get_connection()
    c = conn.cursor()

    c.executescript("""
    CREATE TABLE IF NOT EXISTS operatori_economici (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ragione_sociale TEXT NOT NULL,
        tipo_soggetto TEXT DEFAULT 'Impresa',
        codice_fiscale TEXT,
        partita_iva TEXT,
        indirizzo TEXT,
        cap TEXT,
        citta TEXT,
        provincia TEXT,
        pec TEXT,
        email TEXT,
        telefono TEXT,
        iban_dedicato TEXT,
        legale_rappresentante TEXT,
        attivo INTEGER DEFAULT 1,
        fid_excel TEXT,
        data_creazione TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS affidamenti (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        oggetto TEXT NOT NULL,
        tipo_procedura TEXT,
        tipo_prestazione TEXT,
        classificazione_soglia TEXT,
        cig TEXT,
        cup TEXT,
        id_operatore INTEGER REFERENCES operatori_economici(id),
        id_gara INTEGER REFERENCES gare(id),
        id_contratto INTEGER REFERENCES contratti(id),
        rup TEXT,
        qualifica_rup TEXT,
        dirigente TEXT DEFAULT 'Arch. Linda Peruggi',
        qualifica_dirigente TEXT DEFAULT 'Dirigente del Settore Sviluppo Economico, Ambientale e Floricoltura',
        direttore_lavori TEXT,
        qualifica_dl TEXT,
        collaboratori TEXT,
        capitolo_bilancio TEXT,
        esercizio TEXT,
        rif_normativo TEXT,
        forma_contratto TEXT DEFAULT 'corrispondenza',
        stato TEXT DEFAULT 'Bozza',
        importo_affidato REAL DEFAULT 0,
        premessa_1 TEXT,
        premessa_2 TEXT,
        premessa_3 TEXT,
        prot_preventivo TEXT,
        data_preventivo TEXT,
        prot_durc TEXT,
        validita_durc TEXT,
        tempi_esecuzione TEXT,
        penali TEXT,
        data_creazione TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS quadro_economico (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_affidamento INTEGER NOT NULL REFERENCES affidamenti(id) ON DELETE CASCADE,
        sezione TEXT DEFAULT 'A',
        descrizione TEXT NOT NULL,
        importo REAL DEFAULT 0,
        aliquota_iva REAL DEFAULT 22,
        soggetto_iva INTEGER DEFAULT 1,
        destinazione TEXT DEFAULT 'Fornitore',
        ordine INTEGER DEFAULT 0
    );

    CREATE TABLE IF NOT EXISTS coperture_finanziarie (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_affidamento INTEGER NOT NULL REFERENCES affidamenti(id) ON DELETE CASCADE,
        missione TEXT,
        programma TEXT,
        titolo TEXT,
        macroaggregato TEXT,
        capitolo TEXT,
        nome_capitolo TEXT,
        anno_bilancio TEXT,
        annualita TEXT,
        importo REAL DEFAULT 0,
        note TEXT,
        ordine INTEGER DEFAULT 0
    );

    CREATE TABLE IF NOT EXISTS movimenti_contabili (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_affidamento INTEGER NOT NULL REFERENCES affidamenti(id),
        id_determina INTEGER REFERENCES determine(id),
        id_fattura INTEGER REFERENCES fatture(id),
        tipo_fase TEXT NOT NULL,
        importo REAL DEFAULT 0,
        capitolo_bilancio TEXT,
        numero_impegno TEXT,
        esercizio TEXT,
        stato TEXT DEFAULT 'Attivo',
        data_movimento DATE,
        note TEXT,
        data_creazione TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS preventivi (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_affidamento INTEGER NOT NULL REFERENCES affidamenti(id),
        id_operatore INTEGER NOT NULL REFERENCES operatori_economici(id),
        protocollo TEXT,
        data_protocollo DATE,
        importo_netto REAL DEFAULT 0,
        aliquota_iva REAL DEFAULT 22,
        importo_totale REAL DEFAULT 0,
        selezionato INTEGER DEFAULT 0,
        note TEXT,
        data_creazione TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS determine (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tipo_determina TEXT NOT NULL,
        numero TEXT,
        anno INTEGER,
        data_determina DATE,
        oggetto TEXT,
        id_affidamento INTEGER REFERENCES affidamenti(id),
        id_fattura INTEGER REFERENCES fatture(id),
        id_determina_padre INTEGER REFERENCES determine(id),
        importo REAL DEFAULT 0,
        capitolo_bilancio TEXT,
        impegno_spesa TEXT,
        esercizio TEXT,
        rif_normativo TEXT,
        stato_iter TEXT DEFAULT 'Bozza',
        indicazioni_ai TEXT,
        snapshot_qe TEXT,
        snapshot_coperture TEXT,
        file_path TEXT,
        data_creazione TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS personale (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome_cognome TEXT NOT NULL,
        qualifica TEXT,
        attivo INTEGER DEFAULT 1,
        data_creazione TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS checklist_verifiche (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_affidamento INTEGER NOT NULL REFERENCES affidamenti(id),
        tipo_verifica TEXT NOT NULL,
        esito INTEGER,
        obbligatoria INTEGER DEFAULT 1,
        bloccante INTEGER DEFAULT 1,
        data_verifica DATE,
        protocollo TEXT,
        data_scadenza DATE,
        note TEXT,
        data_creazione TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS varianti (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_affidamento INTEGER NOT NULL REFERENCES affidamenti(id),
        id_determina INTEGER REFERENCES determine(id),
        descrizione TEXT NOT NULL,
        importo_variante REAL DEFAULT 0,
        tipo_variante TEXT DEFAULT 'Integrativa',
        data_approvazione DATE,
        stato TEXT DEFAULT 'Proposta',
        motivazione TEXT,
        data_creazione TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS verbali_dl (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_affidamento INTEGER NOT NULL REFERENCES affidamenti(id),
        tipo_verbale TEXT NOT NULL,
        fase TEXT,
        numero_progressivo INTEGER DEFAULT 1,
        data_verbale DATE,
        redattore TEXT,
        importo_sal REAL,
        note TEXT,
        file_path TEXT,
        data_creazione TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS fatture (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_affidamento INTEGER REFERENCES affidamenti(id),
        id_operatore INTEGER REFERENCES operatori_economici(id),
        numero_fattura TEXT NOT NULL,
        data_fattura DATE,
        importo_netto REAL DEFAULT 0,
        aliquota_iva REAL DEFAULT 22,
        importo_iva REAL DEFAULT 0,
        importo_totale REAL DEFAULT 0,
        protocollo_pec TEXT,
        data_protocollo_pec DATE,
        tipo_liquidazione TEXT DEFAULT 'Saldo',
        stato TEXT DEFAULT 'Registrata',
        stato_pagamento TEXT DEFAULT 'Da liquidare',
        data_creazione TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS pubblicazioni (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_determina INTEGER REFERENCES determine(id),
        id_concessione INTEGER REFERENCES concessioni_demaniali(id),
        canale TEXT NOT NULL,
        stato_pubblicazione TEXT DEFAULT 'Da pubblicare',
        esito TEXT,
        data_pubblicazione DATE,
        data_scadenza DATE,
        numero_registro TEXT,
        data_creazione TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS gare (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tipo_gara TEXT,
        oggetto TEXT NOT NULL,
        cig TEXT,
        cup TEXT,
        criterio_aggiudicazione TEXT DEFAULT 'Offerta economicamente più vantaggiosa',
        rup TEXT,
        importo_base_asta REAL DEFAULT 0,
        piattaforma TEXT,
        scadenza_offerte DATE,
        data_aggiudicazione DATE,
        importo_aggiudicazione REAL,
        ribasso_percentuale REAL,
        stato TEXT DEFAULT 'In preparazione',
        fase_corrente TEXT DEFAULT 'Preparazione',
        note TEXT,
        data_creazione TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS offerte_gara (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_gara INTEGER NOT NULL REFERENCES gare(id),
        id_operatore INTEGER NOT NULL REFERENCES operatori_economici(id),
        importo_offerto REAL,
        ribasso_percentuale REAL,
        punteggio_tecnico REAL,
        punteggio_economico REAL,
        punteggio_totale REAL,
        posizione_graduatoria INTEGER,
        aggiudicatario INTEGER DEFAULT 0,
        esito TEXT
    );

    CREATE TABLE IF NOT EXISTS contratti (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_affidamento INTEGER REFERENCES affidamenti(id),
        id_operatore INTEGER REFERENCES operatori_economici(id),
        numero_repertorio TEXT,
        data_stipula DATE,
        oggetto TEXT,
        importo REAL,
        data_inizio DATE,
        data_fine DATE,
        note TEXT,
        data_creazione TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS coperture_finanziarie (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_affidamento INTEGER NOT NULL REFERENCES affidamenti(id) ON DELETE CASCADE,
        missione TEXT,
        programma TEXT,
        titolo TEXT,
        macroaggregato TEXT,
        capitolo TEXT,
        nome_capitolo TEXT,
        anno_bilancio TEXT,
        annualita TEXT,
        importo REAL DEFAULT 0,
        note TEXT,
        ordine INTEGER DEFAULT 0
    );

    CREATE TABLE IF NOT EXISTS allegati (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        entita_tipo TEXT NOT NULL,
        id_entita INTEGER NOT NULL,
        nome_file TEXT NOT NULL,
        percorso_file TEXT NOT NULL,
        tipo_file TEXT,
        descrizione TEXT,
        data_caricamento TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS log_attivita (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        entita_tipo TEXT NOT NULL,
        id_entita INTEGER NOT NULL,
        azione TEXT NOT NULL,
        dettaglio TEXT,
        snapshot_json TEXT,
        utente TEXT DEFAULT 'sistema',
        data_ora TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS concessioni_demaniali (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tipo_concessione TEXT,
        oggetto TEXT NOT NULL,
        id_operatore INTEGER REFERENCES operatori_economici(id),
        numero_concessione TEXT,
        data_rilascio DATE,
        durata_anni INTEGER,
        data_scadenza DATE,
        canone_annuo REAL DEFAULT 0,
        ubicazione TEXT,
        superficie_mq REAL,
        stato TEXT DEFAULT 'Attiva',
        cig TEXT,
        rif_normativo TEXT,
        note TEXT,
        data_creazione TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS impostazioni (
        chiave TEXT PRIMARY KEY,
        valore TEXT
    );
    """)

    # Settore e servizio di default
    c.execute("INSERT OR IGNORE INTO impostazioni VALUES ('settore_nome','Settore Sviluppo Economico, Ambientale e Floricoltura')")
    c.execute("INSERT OR IGNORE INTO impostazioni VALUES ('servizio_nome','Servizio Demanio Marittimo')")
    c.execute("INSERT OR IGNORE INTO impostazioni VALUES ('dirigente_default','Arch. Linda Peruggi')")
    c.execute("INSERT OR IGNORE INTO impostazioni VALUES ('qualifica_dirigente_default','Dirigente del Settore Sviluppo Economico, Ambientale e Floricoltura')")

    # Migrazioni per tabelle esistenti
    try:
        c.execute("ALTER TABLE determine ADD COLUMN indicazioni_ai TEXT")
    except sqlite3.OperationalError: pass
    try:
        c.execute("ALTER TABLE determine ADD COLUMN snapshot_qe TEXT")
    except sqlite3.OperationalError: pass
    try:
        c.execute("ALTER TABLE determine ADD COLUMN snapshot_coperture TEXT")
    except sqlite3.OperationalError: pass
    
    try:
        c.execute("ALTER TABLE affidamenti ADD COLUMN direttore_lavori TEXT")
    except sqlite3.OperationalError: pass
    try:
        c.execute("ALTER TABLE affidamenti ADD COLUMN qualifica_dl TEXT")
    except sqlite3.OperationalError: pass
    try:
        c.execute("ALTER TABLE affidamenti ADD COLUMN collaboratori TEXT")
    except sqlite3.OperationalError: pass

    # Inizializza personale se vuoto
    if c.execute("SELECT COUNT(*) FROM personale").fetchone()[0] == 0:
        pers = [
            ("Arch. Linda Peruggi", "Dirigente del Settore Sviluppo Economico, Ambientale e Floricoltura"),
            ("Ing. Danilo Burastero", "Funzionario Tecnico"),
            ("Geom. Sergio D'Ighero", "Istruttore Tecnico"),
        ]
        for p in pers:
            c.execute("INSERT INTO personale (nome_cognome, qualifica) VALUES (?,?)", p)

    conn.commit()
    conn.close()


# Init automatico
init_db()


# ═══════════════════════════════════════════════════════════════
# UTILITY
# ═══════════════════════════════════════════════════════════════

def format_importo(valore):
    if valore is None:
        return "0,00"
    s = f"{valore:,.2f}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")


def format_data(data_str):
    if not data_str:
        return ""
    try:
        return datetime.strptime(data_str, "%Y-%m-%d").strftime("%d/%m/%Y")
    except ValueError:
        return data_str


def classificazione_soglia(tipo_prestazione, importo):
    soglia = SOGLIE_COMUNITARIE.get(tipo_prestazione, 5_382_000)
    soglia_ad = SOGLIE_AFFIDAMENTO_DIRETTO.get(tipo_prestazione, 140_000)
    if importo <= soglia_ad:
        return "Sotto soglia — aff. diretto"
    elif importo <= soglia:
        return "Sotto soglia comunitaria"
    else:
        return "Sopra soglia comunitaria"


def procedura_suggerita(tipo_prestazione, importo):
    soglia_ad = SOGLIE_AFFIDAMENTO_DIRETTO.get(tipo_prestazione, 140_000)
    if tipo_prestazione in ("Concessione demaniale", "Concessione SLA (art. 45-bis Cod. Nav.)"):
        return "Procedura aperta (art. 71)"
    if importo <= soglia_ad:
        if tipo_prestazione == "Lavori":
            return "Affidamento diretto (art. 50 c.1 lett. a)"
        return "Affidamento diretto (art. 50 c.1 lett. b)"
    soglia = SOGLIE_COMUNITARIE.get(tipo_prestazione, 5_382_000)
    if importo <= soglia:
        if tipo_prestazione == "Lavori" and importo <= 1_000_000:
            return "Procedura negoziata senza bando (art. 50 c.1 lett. c)"
        elif tipo_prestazione == "Lavori":
            return "Procedura negoziata senza bando (art. 50 c.1 lett. d)"
        return "Procedura negoziata senza bando (art. 50 c.1 lett. c)"
    return "Procedura aperta (art. 71)"


def _log(conn, entita_tipo, id_entita, azione, dettaglio="", snapshot=None, utente="sistema"):
    conn.execute(
        "INSERT INTO log_attivita (entita_tipo, id_entita, azione, dettaglio, snapshot_json, utente) VALUES (?,?,?,?,?,?)",
        (entita_tipo, id_entita, azione, dettaglio, json.dumps(snapshot) if snapshot else None, utente),
    )


# ═══════════════════════════════════════════════════════════════
# OPERATORI ECONOMICI (ex Fornitori)
# ═══════════════════════════════════════════════════════════════

def inserisci_operatore(data):
    with get_db() as conn:
        c = conn.execute(
            """INSERT INTO operatori_economici (ragione_sociale, tipo_soggetto, codice_fiscale,
            partita_iva, indirizzo, cap, citta, provincia, pec, email, telefono,
            iban_dedicato, legale_rappresentante, fid_excel)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (data.get("ragione_sociale"), data.get("tipo_soggetto", "Impresa"),
             data.get("codice_fiscale"), data.get("partita_iva"),
             data.get("indirizzo"), data.get("cap"), data.get("citta"), data.get("provincia"),
             data.get("pec"), data.get("email"), data.get("telefono"),
             data.get("iban_dedicato"), data.get("legale_rappresentante"), data.get("fid_excel")),
        )
        _log(conn, "operatori_economici", c.lastrowid, "inserimento", data.get("ragione_sociale"))
        return c.lastrowid


def aggiorna_operatore(oid, data):
    with get_db() as conn:
        conn.execute(
            """UPDATE operatori_economici SET ragione_sociale=?, tipo_soggetto=?, codice_fiscale=?,
            partita_iva=?, indirizzo=?, cap=?, citta=?, provincia=?, pec=?, email=?, telefono=?,
            iban_dedicato=?, legale_rappresentante=?, fid_excel=? WHERE id=?""",
            (data.get("ragione_sociale"), data.get("tipo_soggetto"),
             data.get("codice_fiscale"), data.get("partita_iva"),
             data.get("indirizzo"), data.get("cap"), data.get("citta"), data.get("provincia"),
             data.get("pec"), data.get("email"), data.get("telefono"),
             data.get("iban_dedicato"), data.get("legale_rappresentante"),
             data.get("fid_excel"), oid),
        )
        _log(conn, "operatori_economici", oid, "aggiornamento")


def get_operatore(oid):
    conn = get_connection()
    r = conn.execute("SELECT * FROM operatori_economici WHERE id=?", (oid,)).fetchone()
    conn.close()
    return r


def cerca_operatori(search=""):
    conn = get_connection()
    if search:
        rows = conn.execute(
            """SELECT * FROM operatori_economici WHERE attivo=1 AND
            (ragione_sociale LIKE ? OR partita_iva LIKE ? OR codice_fiscale LIKE ?)
            ORDER BY ragione_sociale""",
            (f"%{search}%",) * 3,
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM operatori_economici WHERE attivo=1 ORDER BY ragione_sociale").fetchall()
    conn.close()
    return rows


# ═══════════════════════════════════════════════════════════════
# PERSONALE (Anagrafe soggetti interni)
# ═══════════════════════════════════════════════════════════════

def get_personale(attivi_only=False):
    conn = get_connection()
    q = "SELECT * FROM personale"
    if attivi_only: q += " WHERE attivo=1"
    q += " ORDER BY nome_cognome"
    rows = conn.execute(q).fetchall()
    conn.close()
    return rows

def inserisci_personale(data):
    with get_db() as conn:
        c = conn.execute(
            "INSERT INTO personale (nome_cognome, qualifica, attivo) VALUES (?,?,?)",
            (data.get("nome_cognome"), data.get("qualifica"), data.get("attivo", 1))
        )
        return c.lastrowid

def aggiorna_personale(pid, data):
    with get_db() as conn:
        conn.execute(
            "UPDATE personale SET nome_cognome=?, qualifica=?, attivo=? WHERE id=?",
            (data.get("nome_cognome"), data.get("qualifica"), data.get("attivo", 1), pid)
        )

def elimina_personale(pid):
    with get_db() as conn:
        conn.execute("DELETE FROM personale WHERE id=?", (pid,))


# ═══════════════════════════════════════════════════════════════
# AFFIDAMENTI
# ═══════════════════════════════════════════════════════════════

def inserisci_affidamento(data):
    with get_db() as conn:
        c = conn.execute(
            """INSERT INTO affidamenti (oggetto, tipo_procedura, tipo_prestazione,
            classificazione_soglia, cig, cup, id_operatore, id_gara, id_contratto,
            rup, qualifica_rup, dirigente, qualifica_dirigente,
            direttore_lavori, qualifica_dl, collaboratori,
            capitolo_bilancio, esercizio, rif_normativo, forma_contratto,
            stato, importo_affidato,
            premessa_1, premessa_2, premessa_3,
            prot_preventivo, data_preventivo, prot_durc, validita_durc,
            tempi_esecuzione, penali)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (data.get("oggetto"), data.get("tipo_procedura"), data.get("tipo_prestazione"),
             data.get("classificazione_soglia"), data.get("cig"), data.get("cup"),
             data.get("id_operatore"), data.get("id_gara"), data.get("id_contratto"),
             data.get("rup"), data.get("qualifica_rup"),
             data.get("dirigente"),
             data.get("qualifica_dirigente"),
             data.get("direttore_lavori"), data.get("qualifica_dl"), data.get("collaboratori"),
             data.get("capitolo_bilancio"), data.get("esercizio"),
             data.get("rif_normativo"), data.get("forma_contratto", "corrispondenza"),
             data.get("stato", "Bozza"), data.get("importo_affidato", 0),
             data.get("premessa_1"), data.get("premessa_2"), data.get("premessa_3"),
             data.get("prot_preventivo"), data.get("data_preventivo"),
             data.get("prot_durc"), data.get("validita_durc"),
             data.get("tempi_esecuzione"), data.get("penali")),
        )
        _log(conn, "affidamenti", c.lastrowid, "inserimento", data.get("oggetto"))
        return c.lastrowid


def aggiorna_affidamento(aid, data):
    with get_db() as conn:
        # Snapshot prima della modifica
        old = dict(conn.execute("SELECT * FROM affidamenti WHERE id=?", (aid,)).fetchone())
        conn.execute(
            """UPDATE affidamenti SET oggetto=?, tipo_procedura=?, tipo_prestazione=?,
            classificazione_soglia=?, cig=?, cup=?, id_operatore=?, id_gara=?, id_contratto=?,
            rup=?, qualifica_rup=?, dirigente=?, qualifica_dirigente=?,
            direttore_lavori=?, qualifica_dl=?, collaboratori=?,
            capitolo_bilancio=?, esercizio=?, rif_normativo=?, forma_contratto=?,
            stato=?, importo_affidato=?,
            premessa_1=?, premessa_2=?, premessa_3=?,
            prot_preventivo=?, data_preventivo=?, prot_durc=?, validita_durc=?,
            tempi_esecuzione=?, penali=? WHERE id=?""",
            (data.get("oggetto"), data.get("tipo_procedura"), data.get("tipo_prestazione"),
             data.get("classificazione_soglia"), data.get("cig"), data.get("cup"),
             data.get("id_operatore"), data.get("id_gara"), data.get("id_contratto"),
             data.get("rup"), data.get("qualifica_rup"),
             data.get("dirigente"), data.get("qualifica_dirigente"),
             data.get("direttore_lavori"), data.get("qualifica_dl"), data.get("collaboratori"),
             data.get("capitolo_bilancio"), data.get("esercizio"),
             data.get("rif_normativo"), data.get("forma_contratto"),
             data.get("stato"), data.get("importo_affidato", 0),
             data.get("premessa_1"), data.get("premessa_2"), data.get("premessa_3"),
             data.get("prot_preventivo"), data.get("data_preventivo"),
             data.get("prot_durc"), data.get("validita_durc"),
             data.get("tempi_esecuzione"), data.get("penali"), aid),
        )
        _log(conn, "affidamenti", aid, "aggiornamento", snapshot={"prima": old})


def get_affidamento(aid):
    conn = get_connection()
    r = conn.execute(
        """SELECT a.*, o.ragione_sociale as operatore_nome,
        o.partita_iva as operatore_piva, o.codice_fiscale as operatore_cf,
        o.indirizzo as operatore_indirizzo, o.citta as operatore_citta,
        o.provincia as operatore_provincia, o.cap as operatore_cap,
        o.pec as operatore_pec, o.iban_dedicato as operatore_iban,
        o.legale_rappresentante as operatore_legale_rappr
        FROM affidamenti a
        LEFT JOIN operatori_economici o ON a.id_operatore=o.id
        WHERE a.id=?""",
        (aid,),
    ).fetchone()
    conn.close()
    return r


def cerca_affidamenti(search="", tipo="", stato=""):
    conn = get_connection()
    sql = """SELECT a.*, o.ragione_sociale as operatore_nome
        FROM affidamenti a LEFT JOIN operatori_economici o ON a.id_operatore=o.id WHERE 1=1"""
    params = []
    if search:
        sql += " AND (a.oggetto LIKE ? OR a.cig LIKE ? OR o.ragione_sociale LIKE ?)"
        params += [f"%{search}%"] * 3
    if tipo:
        sql += " AND a.tipo_prestazione=?"
        params.append(tipo)
    if stato:
        sql += " AND a.stato=?"
        params.append(stato)
    sql += " ORDER BY a.id DESC"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return rows


def cambia_stato_affidamento(aid, nuovo_stato, utente="sistema"):
    """Cambia stato con validazione della macchina a stati."""
    with get_db() as conn:
        r = conn.execute("SELECT stato FROM affidamenti WHERE id=?", (aid,)).fetchone()
        if not r:
            raise ValueError(f"Affidamento {aid} non trovato")
        stato_corrente = r["stato"]
        validi = TRANSIZIONI_STATO.get(stato_corrente, [])
        if nuovo_stato not in validi:
            raise ValueError(
                f"Transizione non valida: {stato_corrente} → {nuovo_stato}. "
                f"Stati possibili: {', '.join(validi)}"
            )
        conn.execute("UPDATE affidamenti SET stato=? WHERE id=?", (nuovo_stato, aid))
        _log(conn, "affidamenti", aid, "cambio_stato",
             f"{stato_corrente} → {nuovo_stato}", utente=utente)


def elimina_affidamento(aid):
    with get_db() as conn:
        conn.execute("DELETE FROM quadro_economico WHERE id_affidamento=?", (aid,))
        conn.execute("DELETE FROM coperture_finanziarie WHERE id_affidamento=?", (aid,))
        conn.execute("DELETE FROM affidamenti WHERE id=?", (aid,))
        _log(conn, "affidamenti", aid, "eliminazione")


# ── Campi calcolati ──

def get_importi_affidamento(aid):
    """Calcola importi derivati: varianti, autorizzato, liquidato, residuo."""
    conn = get_connection()
    r = conn.execute("SELECT importo_affidato FROM affidamenti WHERE id=?", (aid,)).fetchone()
    affidato = r["importo_affidato"] if r else 0

    varianti = conn.execute(
        "SELECT COALESCE(SUM(importo_variante),0) as tot FROM varianti WHERE id_affidamento=? AND stato='Approvata'",
        (aid,),
    ).fetchone()["tot"]

    liquidato = conn.execute(
        "SELECT COALESCE(SUM(importo),0) as tot FROM movimenti_contabili WHERE id_affidamento=? AND tipo_fase='Liquidazione' AND stato='Attivo'",
        (aid,),
    ).fetchone()["tot"]

    autorizzato = affidato + varianti
    residuo = autorizzato - liquidato
    conn.close()
    return {
        "importo_affidato": affidato,
        "importo_varianti": varianti,
        "importo_totale_autorizzato": autorizzato,
        "totale_liquidato": liquidato,
        "residuo": residuo,
    }


# ═══════════════════════════════════════════════════════════════
# QUADRO ECONOMICO
# ═══════════════════════════════════════════════════════════════

def get_quadro_economico(aid):
    conn = get_connection()
    rows = conn.execute("SELECT * FROM quadro_economico WHERE id_affidamento=? ORDER BY ordine", (aid,)).fetchall()
    conn.close()
    return rows


def salva_quadro_economico(aid, voci):
    with get_db() as conn:
        conn.execute("DELETE FROM quadro_economico WHERE id_affidamento=?", (aid,))
        for v in voci:
            conn.execute(
                """INSERT INTO quadro_economico (id_affidamento, sezione, descrizione, importo,
                aliquota_iva, soggetto_iva, destinazione, ordine) VALUES (?,?,?,?,?,?,?,?)""",
                (aid, v.get("sezione", "A"), v.get("descrizione"), v.get("importo", 0),
                 v.get("aliquota_iva", 22), v.get("soggetto_iva", 1),
                 v.get("destinazione", "Fornitore"), v.get("ordine", 0)),
            )


# ═══════════════════════════════════════════════════════════════
# COPERTURE FINANZIARIE (capitoli di bilancio multipli)
# ═══════════════════════════════════════════════════════════════

def get_coperture(aid):
    conn = get_connection()
    rows = conn.execute("SELECT * FROM coperture_finanziarie WHERE id_affidamento=? ORDER BY ordine", (aid,)).fetchall()
    conn.close()
    return rows


def salva_coperture(aid, voci):
    with get_db() as conn:
        conn.execute("DELETE FROM coperture_finanziarie WHERE id_affidamento=?", (aid,))
        for v in voci:
            conn.execute(
                """INSERT INTO coperture_finanziarie (id_affidamento, missione, programma, titolo,
                macroaggregato, capitolo, nome_capitolo, anno_bilancio, annualita, importo, note, ordine)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                (aid, v.get("missione"), v.get("programma"), v.get("titolo"),
                 v.get("macroaggregato"), v.get("capitolo"), v.get("nome_capitolo"),
                 v.get("anno_bilancio"), v.get("annualita"), v.get("importo", 0),
                 v.get("note"), v.get("ordine", 0)),
            )


def format_copertura_breve(cop):
    """Formatta una copertura per il documento: Miss.X/Prog.Y/Tit.Z/Cap.N"""
    parts = []
    if cop.get("missione"): parts.append(f"Miss. {cop['missione']}")
    if cop.get("programma"): parts.append(f"Prog. {cop['programma']}")
    if cop.get("titolo"): parts.append(f"Tit. {cop['titolo']}")
    if cop.get("macroaggregato"): parts.append(f"Macr. {cop['macroaggregato']}")
    if cop.get("capitolo"): parts.append(f"Cap. {cop['capitolo']}")
    return " / ".join(parts) if parts else ""


# ═══════════════════════════════════════════════════════════════
# DETERMINE
# ═══════════════════════════════════════════════════════════════

def inserisci_determina(data):
    with get_db() as conn:
        c = conn.execute(
            """INSERT INTO determine (tipo_determina, numero, anno, data_determina, oggetto,
            id_affidamento, id_fattura, id_determina_padre, importo,
            capitolo_bilancio, impegno_spesa, esercizio, rif_normativo, stato_iter, 
            indicazioni_ai, snapshot_qe, snapshot_coperture)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (data.get("tipo_determina"), data.get("numero"), data.get("anno"),
             data.get("data_determina"), data.get("oggetto"),
             data.get("id_affidamento"), data.get("id_fattura"), data.get("id_determina_padre"),
             data.get("importo", 0), data.get("capitolo_bilancio"),
             data.get("impegno_spesa"), data.get("esercizio"),
             data.get("rif_normativo"), data.get("stato_iter", "Bozza"),
             data.get("indicazioni_ai"), data.get("snapshot_qe"), data.get("snapshot_coperture")),
        )
        _log(conn, "determine", c.lastrowid, "inserimento", data.get("tipo_determina"))
        return c.lastrowid


def aggiorna_determina(did, data):
    with get_db() as conn:
        conn.execute(
            """UPDATE determine SET tipo_determina=?, numero=?, anno=?, data_determina=?, oggetto=?,
            id_affidamento=?, id_fattura=?, id_determina_padre=?, importo=?,
            capitolo_bilancio=?, impegno_spesa=?, esercizio=?, rif_normativo=?,
            stato_iter=?, indicazioni_ai=?, snapshot_qe=?, snapshot_coperture=?, file_path=? WHERE id=?""",
            (data.get("tipo_determina"), data.get("numero"), data.get("anno"),
             data.get("data_determina"), data.get("oggetto"),
             data.get("id_affidamento"), data.get("id_fattura"), data.get("id_determina_padre"),
             data.get("importo", 0), data.get("capitolo_bilancio"),
             data.get("impegno_spesa"), data.get("esercizio"),
             data.get("rif_normativo"), data.get("stato_iter"),
             data.get("indicazioni_ai"), data.get("snapshot_qe"), data.get("snapshot_coperture"),
             data.get("file_path"), did),
        )
        _log(conn, "determine", did, "aggiornamento")


def elimina_determina(did):
    with get_db() as conn:
        conn.execute("DELETE FROM pubblicazioni WHERE id_determina=?", (did,))
        conn.execute("DELETE FROM movimenti_contabili WHERE id_determina=?", (did,))
        conn.execute("DELETE FROM determine WHERE id=?", (did,))
        _log(conn, "determine", did, "eliminazione")


def get_determina(did):
    conn = get_connection()
    r = conn.execute(
        """SELECT d.*,
        a.oggetto as affidamento_oggetto, a.cig, a.cup, a.rup,
        a.capitolo_bilancio as aff_capitolo, a.rif_normativo as aff_rif_normativo,
        a.dirigente, a.qualifica_dirigente, a.importo_affidato,
        a.tipo_prestazione, a.tipo_procedura,
        o.ragione_sociale, o.partita_iva, o.codice_fiscale,
        o.indirizzo as fornitore_indirizzo, o.citta as fornitore_citta,
        o.provincia as fornitore_provincia, o.cap as fornitore_cap,
        o.pec as fornitore_pec, o.iban_dedicato,
        o.legale_rappresentante,
        f.numero_fattura, f.data_fattura, f.importo_netto as fat_netto,
        f.aliquota_iva as fat_aliquota, f.importo_iva as fat_iva,
        f.importo_totale as fat_totale,
        dp.tipo_determina as padre_tipo, dp.numero as padre_numero,
        dp.data_determina as padre_data
        FROM determine d
        LEFT JOIN affidamenti a ON d.id_affidamento=a.id
        LEFT JOIN operatori_economici o ON a.id_operatore=o.id
        LEFT JOIN fatture f ON d.id_fattura=f.id
        LEFT JOIN determine dp ON d.id_determina_padre=dp.id
        WHERE d.id=?""",
        (did,),
    ).fetchone()
    conn.close()
    return r


def cerca_determine(tipo="", stato="", id_affidamento=None):
    conn = get_connection()
    sql = """SELECT d.*, a.oggetto as affidamento_oggetto, o.ragione_sociale as operatore_nome,
        dp.tipo_determina as padre_tipo, dp.numero as padre_numero
        FROM determine d
        LEFT JOIN affidamenti a ON d.id_affidamento=a.id
        LEFT JOIN operatori_economici o ON a.id_operatore=o.id
        LEFT JOIN determine dp ON d.id_determina_padre=dp.id WHERE 1=1"""
    params = []
    if tipo:
        sql += " AND d.tipo_determina=?"; params.append(tipo)
    if stato:
        sql += " AND d.stato_iter=?"; params.append(stato)
    if id_affidamento:
        sql += " AND d.id_affidamento=?"; params.append(id_affidamento)
    sql += " ORDER BY d.id DESC"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return rows


def get_iter_affidamento(aid):
    """Restituisce tutte le determine di un affidamento in ordine cronologico."""
    conn = get_connection()
    rows = conn.execute(
        """SELECT d.*, dp.tipo_determina as padre_tipo, dp.numero as padre_numero
        FROM determine d LEFT JOIN determine dp ON d.id_determina_padre=dp.id
        WHERE d.id_affidamento=? ORDER BY d.id""", (aid,)).fetchall()
    conn.close()
    return rows


# ═══════════════════════════════════════════════════════════════
# MOVIMENTI CONTABILI
# ═══════════════════════════════════════════════════════════════

def inserisci_movimento(data):
    with get_db() as conn:
        c = conn.execute(
            """INSERT INTO movimenti_contabili (id_affidamento, id_determina, id_fattura,
            tipo_fase, importo, capitolo_bilancio, numero_impegno, esercizio, stato, data_movimento, note)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (data.get("id_affidamento"), data.get("id_determina"), data.get("id_fattura"),
             data.get("tipo_fase"), data.get("importo", 0),
             data.get("capitolo_bilancio"), data.get("numero_impegno"),
             data.get("esercizio"), data.get("stato", "Attivo"),
             data.get("data_movimento"), data.get("note")),
        )
        _log(conn, "movimenti_contabili", c.lastrowid, "inserimento", data.get("tipo_fase"))
        return c.lastrowid


def get_movimenti(aid):
    conn = get_connection()
    rows = conn.execute(
        "SELECT * FROM movimenti_contabili WHERE id_affidamento=? ORDER BY data_movimento, id", (aid,)).fetchall()
    conn.close()
    return rows


def ha_prenotazione(aid):
    conn = get_connection()
    r = conn.execute(
        "SELECT COUNT(*) as n FROM movimenti_contabili WHERE id_affidamento=? AND tipo_fase='Prenotazione' AND stato='Attivo'",
        (aid,)).fetchone()
    conn.close()
    return r["n"] > 0


def ha_impegno(aid):
    conn = get_connection()
    r = conn.execute(
        "SELECT COUNT(*) as n FROM movimenti_contabili WHERE id_affidamento=? AND tipo_fase='Impegno' AND stato='Attivo'",
        (aid,)).fetchone()
    conn.close()
    return r["n"] > 0


# ═══════════════════════════════════════════════════════════════
# FATTURE
# ═══════════════════════════════════════════════════════════════

def inserisci_fattura(data):
    with get_db() as conn:
        c = conn.execute(
            """INSERT INTO fatture (id_affidamento, id_operatore, numero_fattura, data_fattura,
            importo_netto, aliquota_iva, importo_iva, importo_totale,
            protocollo_pec, data_protocollo_pec, tipo_liquidazione, stato, stato_pagamento)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (data.get("id_affidamento"), data.get("id_operatore"),
             data.get("numero_fattura"), data.get("data_fattura"),
             data.get("importo_netto", 0), data.get("aliquota_iva", 22),
             data.get("importo_iva"), data.get("importo_totale"),
             data.get("protocollo_pec"), data.get("data_protocollo_pec"),
             data.get("tipo_liquidazione", "Saldo"),
             data.get("stato", "Registrata"), data.get("stato_pagamento", "Da liquidare")),
        )
        _log(conn, "fatture", c.lastrowid, "inserimento", data.get("numero_fattura"))
        return c.lastrowid


def aggiorna_fattura(fid, data):
    with get_db() as conn:
        conn.execute(
            """UPDATE fatture SET id_affidamento=?, id_operatore=?, numero_fattura=?, data_fattura=?,
            importo_netto=?, aliquota_iva=?, importo_iva=?, importo_totale=?,
            protocollo_pec=?, data_protocollo_pec=?, tipo_liquidazione=?,
            stato=?, stato_pagamento=? WHERE id=?""",
            (data.get("id_affidamento"), data.get("id_operatore"),
             data.get("numero_fattura"), data.get("data_fattura"),
             data.get("importo_netto", 0), data.get("aliquota_iva", 22),
             data.get("importo_iva"), data.get("importo_totale"),
             data.get("protocollo_pec"), data.get("data_protocollo_pec"),
             data.get("tipo_liquidazione"),
             data.get("stato"), data.get("stato_pagamento"), fid),
        )
        _log(conn, "fatture", fid, "aggiornamento")


def get_fattura(fid):
    conn = get_connection()
    r = conn.execute(
        """SELECT f.*, a.oggetto as affidamento_oggetto, a.cig, a.rup,
        a.rif_normativo, a.dirigente, a.qualifica_dirigente,
        a.capitolo_bilancio, a.esercizio, a.tipo_prestazione,
        o.ragione_sociale, o.partita_iva, o.codice_fiscale,
        o.indirizzo as fornitore_indirizzo, o.citta as fornitore_citta,
        o.iban_dedicato, o.legale_rappresentante
        FROM fatture f
        LEFT JOIN affidamenti a ON f.id_affidamento=a.id
        LEFT JOIN operatori_economici o ON f.id_operatore=o.id
        WHERE f.id=?""", (fid,)).fetchone()
    conn.close()
    return r


def cerca_fatture(stato_pag="", id_affidamento=None):
    conn = get_connection()
    sql = """SELECT f.*, a.oggetto as affidamento_oggetto, o.ragione_sociale as operatore_nome
        FROM fatture f
        LEFT JOIN affidamenti a ON f.id_affidamento=a.id
        LEFT JOIN operatori_economici o ON f.id_operatore=o.id WHERE 1=1"""
    params = []
    if stato_pag:
        sql += " AND f.stato_pagamento=?"; params.append(stato_pag)
    if id_affidamento:
        sql += " AND f.id_affidamento=?"; params.append(id_affidamento)
    sql += " ORDER BY f.data_fattura DESC"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return rows


def elimina_fattura(fid):
    with get_db() as conn:
        conn.execute("DELETE FROM fatture WHERE id=?", (fid,))
        _log(conn, "fatture", fid, "eliminazione")


# ═══════════════════════════════════════════════════════════════
# PREVENTIVI
# ═══════════════════════════════════════════════════════════════

def inserisci_preventivo(data):
    with get_db() as conn:
        c = conn.execute(
            """INSERT INTO preventivi (id_affidamento, id_operatore, protocollo, data_protocollo,
            importo_netto, aliquota_iva, importo_totale, selezionato, note)
            VALUES (?,?,?,?,?,?,?,?,?)""",
            (data.get("id_affidamento"), data.get("id_operatore"),
             data.get("protocollo"), data.get("data_protocollo"),
             data.get("importo_netto", 0), data.get("aliquota_iva", 22),
             data.get("importo_totale", 0), data.get("selezionato", 0), data.get("note")),
        )
        return c.lastrowid


def get_preventivi(aid):
    conn = get_connection()
    rows = conn.execute(
        """SELECT p.*, o.ragione_sociale FROM preventivi p
        LEFT JOIN operatori_economici o ON p.id_operatore=o.id
        WHERE p.id_affidamento=? ORDER BY p.id""", (aid,)).fetchall()
    conn.close()
    return rows


# ═══════════════════════════════════════════════════════════════
# CHECKLIST VERIFICHE
# ═══════════════════════════════════════════════════════════════

def inserisci_verifica(data):
    with get_db() as conn:
        c = conn.execute(
            """INSERT INTO checklist_verifiche (id_affidamento, tipo_verifica, esito,
            obbligatoria, bloccante, data_verifica, protocollo, data_scadenza, note)
            VALUES (?,?,?,?,?,?,?,?,?)""",
            (data.get("id_affidamento"), data.get("tipo_verifica"), data.get("esito"),
             data.get("obbligatoria", 1), data.get("bloccante", 1),
             data.get("data_verifica"), data.get("protocollo"),
             data.get("data_scadenza"), data.get("note")),
        )
        return c.lastrowid


def get_verifiche(aid):
    conn = get_connection()
    rows = conn.execute("SELECT * FROM checklist_verifiche WHERE id_affidamento=? ORDER BY id", (aid,)).fetchall()
    conn.close()
    return rows


def verifiche_ok(aid):
    """True se tutte le verifiche bloccanti hanno esito positivo."""
    conn = get_connection()
    r = conn.execute(
        "SELECT COUNT(*) as n FROM checklist_verifiche WHERE id_affidamento=? AND bloccante=1 AND (esito IS NULL OR esito=0)",
        (aid,)).fetchone()
    conn.close()
    return r["n"] == 0


def elimina_verifica(vid):
    with get_db() as conn:
        conn.execute("DELETE FROM checklist_verifiche WHERE id=?", (vid,))
        _log(conn, "checklist_verifiche", vid, "eliminazione")


def aggiorna_verifica(vid, data):
    with get_db() as conn:
        conn.execute(
            """UPDATE checklist_verifiche SET tipo_verifica=?, esito=?, obbligatoria=?, 
            bloccante=?, data_verifica=?, protocollo=?, data_scadenza=?, note=? WHERE id=?""",
            (data.get("tipo_verifica"), data.get("esito"), data.get("obbligatoria"),
             data.get("bloccante"), data.get("data_verifica"), data.get("protocollo"),
             data.get("data_scadenza"), data.get("note"), vid),
        )
        _log(conn, "checklist_verifiche", vid, "aggiornamento")


# ═══════════════════════════════════════════════════════════════
# VARIANTI
# ═══════════════════════════════════════════════════════════════

def inserisci_variante(data):
    with get_db() as conn:
        c = conn.execute(
            """INSERT INTO varianti (id_affidamento, id_determina, descrizione,
            importo_variante, tipo_variante, data_approvazione, stato, motivazione)
            VALUES (?,?,?,?,?,?,?,?)""",
            (data.get("id_affidamento"), data.get("id_determina"), data.get("descrizione"),
             data.get("importo_variante", 0), data.get("tipo_variante", "Integrativa"),
             data.get("data_approvazione"), data.get("stato", "Proposta"), data.get("motivazione")),
        )
        _log(conn, "varianti", c.lastrowid, "inserimento")
        return c.lastrowid


def get_varianti(aid):
    conn = get_connection()
    rows = conn.execute("SELECT * FROM varianti WHERE id_affidamento=? ORDER BY id", (aid,)).fetchall()
    conn.close()
    return rows


# ═══════════════════════════════════════════════════════════════
# VERBALI DL
# ═══════════════════════════════════════════════════════════════

def inserisci_verbale(data):
    tipo = data.get("tipo_verbale")
    aid = data.get("id_affidamento")
    with get_db() as conn:
        # Vincoli: un solo fine lavori e un solo CRE
        if tipo in ("Fine lavori", "CRE"):
            existing = conn.execute(
                "SELECT COUNT(*) as n FROM verbali_dl WHERE id_affidamento=? AND tipo_verbale=?",
                (aid, tipo)).fetchone()
            if existing["n"] > 0:
                raise ValueError(f"Esiste già un verbale di '{tipo}' per questo affidamento.")

        # Calcola numero progressivo
        last = conn.execute(
            "SELECT MAX(numero_progressivo) as m FROM verbali_dl WHERE id_affidamento=? AND tipo_verbale=?",
            (aid, tipo)).fetchone()
        num = (last["m"] or 0) + 1

        fase = FASI_VERBALE.get(tipo, "esecuzione")
        c = conn.execute(
            """INSERT INTO verbali_dl (id_affidamento, tipo_verbale, fase, numero_progressivo,
            data_verbale, redattore, importo_sal, note)
            VALUES (?,?,?,?,?,?,?,?)""",
            (aid, tipo, fase, num,
             data.get("data_verbale"), data.get("redattore"),
             data.get("importo_sal"), data.get("note")),
        )
        _log(conn, "verbali_dl", c.lastrowid, "inserimento", tipo)
        return c.lastrowid


def get_verbali(aid):
    conn = get_connection()
    rows = conn.execute(
        "SELECT * FROM verbali_dl WHERE id_affidamento=? ORDER BY data_verbale, numero_progressivo", (aid,)).fetchall()
    conn.close()
    return rows


def elimina_verbale(vid):
    with get_db() as conn:
        conn.execute("DELETE FROM verbali_dl WHERE id=?", (vid,))
        _log(conn, "verbali_dl", vid, "eliminazione")


def aggiorna_verbale(vid, data):
    tipo = data.get("tipo_verbale")
    fase = FASI_VERBALE.get(tipo, "esecuzione")
    with get_db() as conn:
        conn.execute(
            """UPDATE verbali_dl SET tipo_verbale=?, fase=?, data_verbale=?, redattore=?, 
            importo_sal=?, note=? WHERE id=?""",
            (tipo, fase, data.get("data_verbale"), data.get("redattore"),
             data.get("importo_sal"), data.get("note"), vid),
        )
        _log(conn, "verbali_dl", vid, "aggiornamento")


# ═══════════════════════════════════════════════════════════════
# GARE
# ═══════════════════════════════════════════════════════════════

def inserisci_gara(data):
    with get_db() as conn:
        c = conn.execute(
            """INSERT INTO gare (tipo_gara, oggetto, cig, cup, criterio_aggiudicazione, rup,
            importo_base_asta, piattaforma, scadenza_offerte,
            data_aggiudicazione, importo_aggiudicazione, ribasso_percentuale,
            stato, fase_corrente, note)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (data.get("tipo_gara"), data.get("oggetto"), data.get("cig"), data.get("cup"),
             data.get("criterio_aggiudicazione", "Offerta economicamente più vantaggiosa"),
             data.get("rup"), data.get("importo_base_asta", 0),
             data.get("piattaforma"), data.get("scadenza_offerte"),
             data.get("data_aggiudicazione"), data.get("importo_aggiudicazione"),
             data.get("ribasso_percentuale"),
             data.get("stato", "In preparazione"), data.get("fase_corrente", "Preparazione"),
             data.get("note")),
        )
        return c.lastrowid


def aggiorna_gara(gid, data):
    with get_db() as conn:
        conn.execute(
            """UPDATE gare SET tipo_gara=?, oggetto=?, cig=?, cup=?, criterio_aggiudicazione=?,
            rup=?, importo_base_asta=?, piattaforma=?, scadenza_offerte=?,
            data_aggiudicazione=?, importo_aggiudicazione=?, ribasso_percentuale=?,
            stato=?, fase_corrente=?, note=? WHERE id=?""",
            (data.get("tipo_gara"), data.get("oggetto"), data.get("cig"), data.get("cup"),
             data.get("criterio_aggiudicazione"), data.get("rup"),
             data.get("importo_base_asta", 0), data.get("piattaforma"),
             data.get("scadenza_offerte"), data.get("data_aggiudicazione"),
             data.get("importo_aggiudicazione"), data.get("ribasso_percentuale"),
             data.get("stato"), data.get("fase_corrente"), data.get("note"), gid),
        )


def get_gara(gid):
    conn = get_connection()
    r = conn.execute("SELECT * FROM gare WHERE id=?", (gid,)).fetchone()
    conn.close()
    return r


def cerca_gare(search="", stato="", fase=""):
    conn = get_connection()
    sql = "SELECT * FROM gare WHERE 1=1"
    params = []
    if search:
        sql += " AND (oggetto LIKE ? OR cig LIKE ?)"; params += [f"%{search}%"] * 2
    if stato:
        sql += " AND stato=?"; params.append(stato)
    if fase:
        sql += " AND fase_corrente=?"; params.append(fase)
    sql += " ORDER BY id DESC"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return rows


def elimina_gara(gid):
    with get_db() as conn:
        conn.execute("DELETE FROM offerte_gara WHERE id_gara=?", (gid,))
        conn.execute("DELETE FROM gare WHERE id=?", (gid,))
        _log(conn, "gare", gid, "eliminazione")


# ═══════════════════════════════════════════════════════════════
# OFFERTE GARA
# ═══════════════════════════════════════════════════════════════

def inserisci_offerta(data):
    with get_db() as conn:
        c = conn.execute(
            """INSERT INTO offerte_gara (id_gara, id_operatore, importo_offerto, ribasso_percentuale,
            punteggio_tecnico, punteggio_economico, punteggio_totale,
            posizione_graduatoria, aggiudicatario, esito)
            VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (data.get("id_gara"), data.get("id_operatore"),
             data.get("importo_offerto"), data.get("ribasso_percentuale"),
             data.get("punteggio_tecnico"), data.get("punteggio_economico"),
             data.get("punteggio_totale"), data.get("posizione_graduatoria"),
             data.get("aggiudicatario", 0), data.get("esito")),
        )
        return c.lastrowid


def get_offerte(gid):
    conn = get_connection()
    rows = conn.execute(
        """SELECT og.*, o.ragione_sociale FROM offerte_gara og
        LEFT JOIN operatori_economici o ON og.id_operatore=o.id
        WHERE og.id_gara=? ORDER BY og.posizione_graduatoria""", (gid,)).fetchall()
    conn.close()
    return rows


# ═══════════════════════════════════════════════════════════════
# CONTRATTI
# ═══════════════════════════════════════════════════════════════

def inserisci_contratto(data):
    with get_db() as conn:
        c = conn.execute(
            """INSERT INTO contratti (id_affidamento, id_operatore, numero_repertorio,
            data_stipula, oggetto, importo, data_inizio, data_fine, note)
            VALUES (?,?,?,?,?,?,?,?,?)""",
            (data.get("id_affidamento"), data.get("id_operatore"),
             data.get("numero_repertorio"), data.get("data_stipula"),
             data.get("oggetto"), data.get("importo"),
             data.get("data_inizio"), data.get("data_fine"), data.get("note")),
        )
        return c.lastrowid


# ═══════════════════════════════════════════════════════════════
# ALLEGATI (polimorfici)
# ═══════════════════════════════════════════════════════════════

def inserisci_allegato(entita_tipo, id_entita, nome_file, percorso_file, tipo_file="", descrizione=""):
    with get_db() as conn:
        c = conn.execute(
            """INSERT INTO allegati (entita_tipo, id_entita, nome_file, percorso_file, tipo_file, descrizione)
            VALUES (?,?,?,?,?,?)""",
            (entita_tipo, id_entita, nome_file, percorso_file, tipo_file, descrizione),
        )
        return c.lastrowid


def get_allegati(entita_tipo, id_entita):
    conn = get_connection()
    rows = conn.execute(
        "SELECT * FROM allegati WHERE entita_tipo=? AND id_entita=? ORDER BY data_caricamento DESC",
        (entita_tipo, id_entita)).fetchall()
    conn.close()
    return rows


# ═══════════════════════════════════════════════════════════════
# PUBBLICAZIONI
# ═══════════════════════════════════════════════════════════════

def inserisci_pubblicazione(data):
    with get_db() as conn:
        c = conn.execute(
            """INSERT INTO pubblicazioni (id_determina, id_concessione, canale,
            stato_pubblicazione, esito, data_pubblicazione, data_scadenza, numero_registro)
            VALUES (?,?,?,?,?,?,?,?)""",
            (data.get("id_determina"), data.get("id_concessione"), data.get("canale"),
             data.get("stato_pubblicazione", "Da pubblicare"), data.get("esito"),
             data.get("data_pubblicazione"), data.get("data_scadenza"),
             data.get("numero_registro")),
        )
        return c.lastrowid


def get_pubblicazioni(id_determina=None, id_concessione=None):
    conn = get_connection()
    if id_determina:
        rows = conn.execute("SELECT * FROM pubblicazioni WHERE id_determina=? ORDER BY canale", (id_determina,)).fetchall()
    elif id_concessione:
        rows = conn.execute("SELECT * FROM pubblicazioni WHERE id_concessione=? ORDER BY canale", (id_concessione,)).fetchall()
    else:
        rows = conn.execute("SELECT * FROM pubblicazioni ORDER BY data_pubblicazione DESC").fetchall()
    conn.close()
    return rows


# ═══════════════════════════════════════════════════════════════
# CONCESSIONI DEMANIALI
# ═══════════════════════════════════════════════════════════════

def inserisci_concessione(data):
    with get_db() as conn:
        c = conn.execute(
            """INSERT INTO concessioni_demaniali (tipo_concessione, oggetto, id_operatore,
            numero_concessione, data_rilascio, durata_anni, data_scadenza,
            canone_annuo, ubicazione, superficie_mq, stato, cig, rif_normativo, note)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (data.get("tipo_concessione"), data.get("oggetto"), data.get("id_operatore"),
             data.get("numero_concessione"), data.get("data_rilascio"),
             data.get("durata_anni"), data.get("data_scadenza"),
             data.get("canone_annuo", 0), data.get("ubicazione"),
             data.get("superficie_mq"), data.get("stato", "Attiva"),
             data.get("cig"), data.get("rif_normativo"), data.get("note")),
        )
        return c.lastrowid


def aggiorna_concessione(cid, data):
    with get_db() as conn:
        conn.execute(
            """UPDATE concessioni_demaniali SET tipo_concessione=?, oggetto=?, id_operatore=?,
            numero_concessione=?, data_rilascio=?, durata_anni=?, data_scadenza=?,
            canone_annuo=?, ubicazione=?, superficie_mq=?, stato=?,
            cig=?, rif_normativo=?, note=? WHERE id=?""",
            (data.get("tipo_concessione"), data.get("oggetto"), data.get("id_operatore"),
             data.get("numero_concessione"), data.get("data_rilascio"),
             data.get("durata_anni"), data.get("data_scadenza"),
             data.get("canone_annuo", 0), data.get("ubicazione"),
             data.get("superficie_mq"), data.get("stato"),
             data.get("cig"), data.get("rif_normativo"), data.get("note"), cid),
        )


def get_concessione(cid):
    conn = get_connection()
    r = conn.execute(
        """SELECT c.*, o.ragione_sociale as operatore_nome
        FROM concessioni_demaniali c
        LEFT JOIN operatori_economici o ON c.id_operatore=o.id
        WHERE c.id=?""", (cid,)).fetchone()
    conn.close()
    return r


def cerca_concessioni(search="", stato=""):
    conn = get_connection()
    sql = """SELECT c.*, o.ragione_sociale as operatore_nome
        FROM concessioni_demaniali c
        LEFT JOIN operatori_economici o ON c.id_operatore=o.id WHERE 1=1"""
    params = []
    if search:
        sql += " AND (c.oggetto LIKE ? OR c.ubicazione LIKE ? OR o.ragione_sociale LIKE ?)"
        params += [f"%{search}%"] * 3
    if stato:
        sql += " AND c.stato=?"; params.append(stato)
    sql += " ORDER BY c.id DESC"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return rows


def elimina_concessione(cid):
    with get_db() as conn:
        conn.execute("DELETE FROM concessioni_demaniali WHERE id=?", (cid,))
        _log(conn, "concessioni_demaniali", cid, "eliminazione")


# ═══════════════════════════════════════════════════════════════
# LOG ATTIVITÀ
# ═══════════════════════════════════════════════════════════════

def get_log(entita_tipo=None, id_entita=None, limit=100):
    conn = get_connection()
    if entita_tipo and id_entita:
        rows = conn.execute(
            "SELECT * FROM log_attivita WHERE entita_tipo=? AND id_entita=? ORDER BY data_ora DESC LIMIT ?",
            (entita_tipo, id_entita, limit)).fetchall()
    else:
        rows = conn.execute("SELECT * FROM log_attivita ORDER BY data_ora DESC LIMIT ?", (limit,)).fetchall()
    conn.close()
    return rows


# ═══════════════════════════════════════════════════════════════
# IMPOSTAZIONI
# ═══════════════════════════════════════════════════════════════

def get_impostazione(chiave, default=""):
    conn = get_connection()
    r = conn.execute("SELECT valore FROM impostazioni WHERE chiave=?", (chiave,)).fetchone()
    conn.close()
    return r["valore"] if r else default


def set_impostazione(chiave, valore):
    with get_db() as conn:
        conn.execute("INSERT OR REPLACE INTO impostazioni (chiave, valore) VALUES (?,?)", (chiave, valore))


def get_all_impostazioni():
    conn = get_connection()
    rows = conn.execute("SELECT * FROM impostazioni").fetchall()
    conn.close()
    return {r["chiave"]: r["valore"] for r in rows}


# ═══════════════════════════════════════════════════════════════
# STATISTICHE DASHBOARD
# ═══════════════════════════════════════════════════════════════

def get_statistiche():
    conn = get_connection()
    st = {}
    st["affidamenti_attivi"] = conn.execute(
        "SELECT COUNT(*) FROM affidamenti WHERE stato NOT IN ('Bozza','Liquidato totale')").fetchone()[0]
    st["fatture_da_liquidare"] = conn.execute(
        "SELECT COUNT(*) FROM fatture WHERE stato_pagamento='Da liquidare'").fetchone()[0]
    st["importo_da_liquidare"] = conn.execute(
        "SELECT COALESCE(SUM(importo_totale),0) FROM fatture WHERE stato_pagamento='Da liquidare'").fetchone()[0]
    st["determine_totali"] = conn.execute("SELECT COUNT(*) FROM determine").fetchone()[0]
    st["gare_attive"] = conn.execute(
        "SELECT COUNT(*) FROM gare WHERE stato NOT IN ('Completata','Annullata')").fetchone()[0]
    st["operatori_totali"] = conn.execute("SELECT COUNT(*) FROM operatori_economici WHERE attivo=1").fetchone()[0]
    st["concessioni_attive"] = conn.execute(
        "SELECT COUNT(*) FROM concessioni_demaniali WHERE stato='Attiva'").fetchone()[0]
    st["verifiche_in_scadenza"] = conn.execute(
        "SELECT COUNT(*) FROM checklist_verifiche WHERE data_scadenza IS NOT NULL AND data_scadenza <= date('now','+30 days') AND (esito IS NULL OR esito=1)").fetchone()[0]
    conn.close()
    return st


# ═══════════════════════════════════════════════════════════════
# REPORT E QUERY
# ═══════════════════════════════════════════════════════════════

def get_nomi_tabelle():
    conn = get_connection()
    rows = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name").fetchall()
    conn.close()
    return [r["name"] for r in rows]


def esegui_report(nome):
    sql = REPORT_PREIMPOSTATI.get(nome)
    if not sql:
        raise ValueError(f"Report '{nome}' non trovato")
    return esegui_query_libera(sql)


def esegui_query_libera(sql):
    sql_upper = sql.strip().upper()
    if not sql_upper.startswith("SELECT"):
        raise ValueError("Solo query SELECT consentite.")
    conn = get_connection()
    c = conn.execute(sql)
    cols = [d[0] for d in c.description] if c.description else []
    rows = c.fetchall()
    conn.close()
    return cols, [tuple(r) for r in rows]


# ═══════════════════════════════════════════════════════════════
# IMPORT EXCEL OPERATORI
# ═══════════════════════════════════════════════════════════════

def get_percorso_excel_fornitori():
    return get_impostazione("excel_fornitori_path")


def salva_percorso_excel_fornitori(path):
    set_impostazione("excel_fornitori_path", path)


def sincronizza_fornitori_excel(excel_path):
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip().lower() for c in ws[1]]

    col_map = {}
    mapping = {
        "ragione_sociale": ["ragione sociale", "nome", "denominazione", "ragione_sociale"],
        "codice_fiscale": ["codice fiscale", "cf", "codice_fiscale"],
        "partita_iva": ["partita iva", "p.iva", "piva", "partita_iva", "p. iva"],
        "indirizzo": ["indirizzo", "via", "sede"],
        "cap": ["cap"],
        "citta": ["città", "citta", "comune"],
        "provincia": ["provincia", "prov"],
        "telefono": ["telefono", "tel"],
        "email": ["email", "e-mail", "mail"],
        "pec": ["pec"],
        "iban_dedicato": ["iban", "iban dedicato", "iban_dedicato"],
        "legale_rappresentante": ["legale rappresentante", "legale_rappresentante", "rappresentante"],
    }
    for field, aliases in mapping.items():
        for i, h in enumerate(headers):
            if h in aliases:
                col_map[field] = i
                break

    if "ragione_sociale" not in col_map:
        raise ValueError("Colonna 'Ragione Sociale' non trovata nel file Excel.")

    result = {"inseriti": 0, "aggiornati": 0, "errori": 0, "dettagli": []}
    with get_db() as conn:
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                nome = str(row[col_map["ragione_sociale"]] or "").strip()
                if not nome:
                    continue
                fid = f"excel_row_{row_idx}"
                data = {"ragione_sociale": nome, "fid_excel": fid}
                for field, col_idx in col_map.items():
                    if field != "ragione_sociale":
                        data[field] = str(row[col_idx] or "").strip() if col_idx < len(row) else ""

                existing = conn.execute("SELECT id FROM operatori_economici WHERE fid_excel=?", (fid,)).fetchone()
                if existing:
                    data_no_fid = {k: v for k, v in data.items() if k != "fid_excel"}
                    sets = ", ".join(f"{k}=?" for k in data_no_fid)
                    conn.execute(f"UPDATE operatori_economici SET {sets} WHERE id=?",
                                 list(data_no_fid.values()) + [existing["id"]])
                    result["aggiornati"] += 1
                else:
                    cols = ", ".join(data.keys())
                    placeholders = ", ".join("?" * len(data))
                    conn.execute(f"INSERT INTO operatori_economici ({cols}) VALUES ({placeholders})", list(data.values()))
                    result["inseriti"] += 1
            except Exception as e:
                result["errori"] += 1
                result["dettagli"].append(f"Errore riga {row_idx}: {e}")

    return result
