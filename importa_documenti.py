"""
Importa Documenti v7.2 — Interazione con documenti esterni.
Gestisce: Excel fornitori, Excel QE, fatture PDF/XML, documenti PDF per contesto AI.
"""

import os
import json
import re


# ═══════════════════════════════════════════════════════════════
# IMPORT QUADRO ECONOMICO DA EXCEL
# ═══════════════════════════════════════════════════════════════

def importa_qe_da_excel(excel_path):
    """
    Legge un foglio Excel e cerca di ricavare le voci del QE.
    Cerca colonne con intestazioni tipo: Descrizione, Importo, Sezione, IVA.
    Ritorna lista di dict compatibili con salva_quadro_economico().
    """
    import openpyxl

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    # Cerca la riga di intestazione (prime 10 righe)
    header_row = None
    col_map = {}
    keywords = {
        "descrizione": ["descrizione", "voce", "lavorazione", "oggetto", "descrizione voce"],
        "importo": ["importo", "ammontare", "costo", "prezzo", "totale"],
        "sezione": ["sezione", "sez", "cat", "categoria"],
        "iva": ["iva", "aliquota", "iva%", "aliquota iva"],
    }

    for row_idx in range(1, min(11, ws.max_row + 1)):
        row_vals = {}
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and isinstance(val, str):
                val_lower = val.strip().lower()
                for key, syns in keywords.items():
                    if any(s in val_lower for s in syns):
                        row_vals[key] = col_idx
        if "descrizione" in row_vals and "importo" in row_vals:
            header_row = row_idx
            col_map = row_vals
            break

    if not header_row:
        # Fallback: colonna A = descrizione, colonna B = importo
        col_map = {"descrizione": 1, "importo": 2}
        header_row = 0  # nessuna intestazione, leggi tutto

    voci = []
    start_row = header_row + 1 if header_row > 0 else 1

    for row_idx in range(start_row, ws.max_row + 1):
        desc = ws.cell(row=row_idx, column=col_map["descrizione"]).value
        imp_cell = ws.cell(row=row_idx, column=col_map["importo"]).value

        if not desc or not str(desc).strip():
            continue
        desc = str(desc).strip()

        # Converti importo
        try:
            if isinstance(imp_cell, (int, float)):
                importo = float(imp_cell)
            elif isinstance(imp_cell, str):
                imp_clean = imp_cell.replace("€", "").replace(".", "").replace(",", ".").strip()
                importo = float(imp_clean) if imp_clean else 0
            else:
                importo = 0
        except (ValueError, TypeError):
            importo = 0

        if importo == 0 and not desc:
            continue

        # Sezione
        sez = "A"
        if "sezione" in col_map:
            sez_val = ws.cell(row=row_idx, column=col_map["sezione"]).value
            if sez_val and "B" in str(sez_val).upper():
                sez = "B"
        # Euristica: "somme a disposizione", "iva", "spese tecniche" → B
        desc_lower = desc.lower()
        if any(k in desc_lower for k in ["somme a disposizione", "spese tecniche", "imprevisti",
                                          "incentivi", "pubblicazione", "contributo anac"]):
            sez = "B"

        # IVA
        aliq = 22
        if "iva" in col_map:
            iva_val = ws.cell(row=row_idx, column=col_map["iva"]).value
            if isinstance(iva_val, (int, float)):
                aliq = float(iva_val)
            elif isinstance(iva_val, str):
                nums = re.findall(r"[\d,\.]+", iva_val)
                if nums:
                    try:
                        aliq = float(nums[0].replace(",", "."))
                    except ValueError:
                        pass

        voci.append({
            "sezione": sez,
            "descrizione": desc,
            "importo": importo,
            "aliquota_iva": aliq,
            "soggetto_iva": 1,
            "destinazione": "Fornitore" if sez == "A" else "Ente",
            "ordine": len(voci),
        })

    wb.close()
    return voci


# ═══════════════════════════════════════════════════════════════
# LETTURA PDF — estrazione testo
# ═══════════════════════════════════════════════════════════════

def estrai_testo_pdf(pdf_path, max_pagine=10):
    """
    Estrae testo da un PDF. Prova prima PyMuPDF (fitz), poi pdfplumber, poi fallback.
    Ritorna il testo come stringa.
    """
    text = ""

    # Tentativo 1: PyMuPDF (fitz) — il più affidabile
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(pdf_path)
        for i, page in enumerate(doc):
            if i >= max_pagine:
                break
            text += page.get_text() + "\n"
        doc.close()
        if text.strip():
            return text.strip()
    except ImportError:
        pass
    except Exception:
        pass

    # Tentativo 2: pdfplumber
    try:
        import pdfplumber
        with pdfplumber.open(pdf_path) as pdf:
            for i, page in enumerate(pdf.pages):
                if i >= max_pagine:
                    break
                t = page.extract_text()
                if t:
                    text += t + "\n"
        if text.strip():
            return text.strip()
    except ImportError:
        pass
    except Exception:
        pass

    # Tentativo 3: pdfminer
    try:
        from pdfminer.high_level import extract_text
        text = extract_text(pdf_path, maxpages=max_pagine)
        if text.strip():
            return text.strip()
    except ImportError:
        pass
    except Exception:
        pass

    if not text.strip():
        raise RuntimeError(
            "Nessuna libreria PDF disponibile.\n"
            "Installa una di queste:\n"
            "  pip install PyMuPDF\n"
            "  pip install pdfplumber\n"
            "  pip install pdfminer.six"
        )

    return text.strip()


# ═══════════════════════════════════════════════════════════════
# IMPORT FATTURA DA PDF/XML (con AI)
# ═══════════════════════════════════════════════════════════════

def importa_fattura_da_file(file_path):
    """
    Legge una fattura da PDF o XML e estrae i dati.
    Per XML (fattura elettronica): parsing diretto.
    Per PDF: usa l'AI per estrarre i campi.
    Ritorna dict con campi fattura, oppure None.
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".xml":
        return _parse_fattura_xml(file_path)
    elif ext == ".pdf":
        return _parse_fattura_pdf(file_path)
    else:
        raise ValueError(f"Formato non supportato: {ext}. Usa PDF o XML.")


def _parse_fattura_xml(xml_path):
    """Parse fattura elettronica XML (formato SDI italiano)."""
    import xml.etree.ElementTree as ET

    tree = ET.parse(xml_path)
    root = tree.getroot()

    # Gestisci namespace
    ns = ""
    if root.tag.startswith("{"):
        ns = root.tag.split("}")[0] + "}"

    def find(path):
        """Cerca con e senza namespace."""
        el = root.find(f".//{ns}{path}")
        if el is None and ns:
            el = root.find(f".//{path}")
        return el.text.strip() if el is not None and el.text else ""

    # Dati cedente (fornitore)
    fornitore = {
        "ragione_sociale": find("CedentePrestatore//Denominazione") or
                           f"{find('CedentePrestatore//Nome')} {find('CedentePrestatore//Cognome')}".strip(),
        "partita_iva": find("CedentePrestatore//IdCodice"),
        "codice_fiscale": find("CedentePrestatore//CodiceFiscale"),
    }

    # Dati fattura
    numero = find("DatiGeneraliDocumento//Numero")
    data = find("DatiGeneraliDocumento//Data")
    importo_totale = find("DatiGeneraliDocumento//ImportoTotaleDocumento")

    # Calcola netto e IVA dai riepiloghi
    netto_totale = 0
    iva_totale = 0
    aliquota = 22

    for riepilogo in root.findall(f".//{ns}DatiRiepilogo") or root.findall(".//DatiRiepilogo"):
        try:
            imp_el = riepilogo.find(f"{ns}ImponibileImporto") or riepilogo.find("ImponibileImporto")
            iva_el = riepilogo.find(f"{ns}Imposta") or riepilogo.find("Imposta")
            aliq_el = riepilogo.find(f"{ns}AliquotaIVA") or riepilogo.find("AliquotaIVA")
            if imp_el is not None:
                netto_totale += float(imp_el.text)
            if iva_el is not None:
                iva_totale += float(iva_el.text)
            if aliq_el is not None:
                aliquota = float(aliq_el.text)
        except (ValueError, TypeError):
            pass

    try:
        imp_tot = float(importo_totale) if importo_totale else (netto_totale + iva_totale)
    except ValueError:
        imp_tot = netto_totale + iva_totale

    return {
        "numero_fattura": numero,
        "data_fattura": data,
        "importo_netto": round(netto_totale, 2),
        "aliquota_iva": aliquota,
        "importo_iva": round(iva_totale, 2),
        "importo_totale": round(imp_tot, 2),
        "fornitore": fornitore,
        "_fonte": "XML fattura elettronica",
    }


def _parse_fattura_pdf(pdf_path):
    """Estrae dati fattura da PDF usando l'AI."""
    testo = estrai_testo_pdf(pdf_path, max_pagine=3)
    if not testo:
        return None

    # Prova prima il parsing con regex (veloce, gratuito)
    result = _parse_fattura_regex(testo)
    if result and result.get("numero_fattura"):
        return result

    # Fallback: usa AI
    return _parse_fattura_con_ai(testo)


def _parse_fattura_regex(testo):
    """Tentativo di parsing fattura con regex — gratuito, funziona su fatture standard."""
    result = {}

    # Numero fattura
    m = re.search(r"(?:fattura|fatt\.?|ft\.?)\s*(?:n\.?|nr\.?|num\.?)\s*[:\s]*([A-Z0-9/\-]+)", testo, re.IGNORECASE)
    if m:
        result["numero_fattura"] = m.group(1).strip()

    # Data fattura
    m = re.search(r"(?:data|del)\s*[:\s]*(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})", testo, re.IGNORECASE)
    if m:
        d = m.group(1).replace(".", "/").replace("-", "/")
        parts = d.split("/")
        if len(parts) == 3:
            if len(parts[2]) == 2:
                parts[2] = "20" + parts[2]
            result["data_fattura"] = f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"

    # Importi
    def find_amount(pattern):
        m = re.search(pattern, testo, re.IGNORECASE)
        if m:
            val = m.group(1).replace(".", "").replace(",", ".").strip()
            try:
                return float(val)
            except ValueError:
                pass
        return None

    result["importo_netto"] = find_amount(r"(?:imponibile|netto|totale\s+imponibile)\s*[:\s€]*\s*([\d.,]+)")
    result["importo_iva"] = find_amount(r"(?:iva|imposta)\s*[:\s€]*\s*([\d.,]+)")
    result["importo_totale"] = find_amount(r"(?:totale\s+(?:fattura|documento|da pagare))\s*[:\s€]*\s*([\d.,]+)")

    # Aliquota IVA
    m = re.search(r"(?:iva|aliquota)\s*[:\s]*(\d{1,2})\s*%", testo, re.IGNORECASE)
    if m:
        result["aliquota_iva"] = float(m.group(1))
    else:
        result["aliquota_iva"] = 22

    # P.IVA fornitore
    m = re.search(r"(?:p\.?\s*iva|partita\s*iva)\s*[:\s]*(\d{11})", testo, re.IGNORECASE)
    if m:
        result["fornitore"] = {"partita_iva": m.group(1)}

    result["_fonte"] = "Parsing PDF (regex)"
    return result if result.get("numero_fattura") else None


def _parse_fattura_con_ai(testo):
    """Usa l'AI per estrarre dati da testo fattura."""
    try:
        from ai_helper import get_api_key, MODEL
        import urllib.request
        import ssl

        api_key = get_api_key()
        if not api_key:
            return None

        prompt = f"""Estrai i dati da questa fattura. Rispondi SOLO con un JSON valido (senza backtick), con questi campi:
- numero_fattura: string
- data_fattura: string formato AAAA-MM-GG
- importo_netto: number
- aliquota_iva: number
- importo_iva: number
- importo_totale: number
- ragione_sociale_fornitore: string
- partita_iva_fornitore: string
- codice_fiscale_fornitore: string

Se un campo non è presente, usa null.

TESTO FATTURA:
{testo[:4000]}"""

        payload = json.dumps({
            "model": MODEL,
            "max_tokens": 500,
            "messages": [{"role": "user", "content": prompt}],
        }).encode("utf-8")

        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=payload,
            headers={
                "Content-Type": "application/json",
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01",
            },
            method="POST",
        )

        ctx = ssl.create_default_context()
        with urllib.request.urlopen(req, timeout=30, context=ctx) as resp:
            data = json.loads(resp.read().decode("utf-8"))

        if data.get("content"):
            text = data["content"][0].get("text", "").strip()
            # Pulisci eventuali backtick
            text = re.sub(r"```json\s*", "", text)
            text = re.sub(r"```\s*", "", text)
            parsed = json.loads(text)
            result = {
                "numero_fattura": parsed.get("numero_fattura"),
                "data_fattura": parsed.get("data_fattura"),
                "importo_netto": parsed.get("importo_netto") or 0,
                "aliquota_iva": parsed.get("aliquota_iva") or 22,
                "importo_iva": parsed.get("importo_iva") or 0,
                "importo_totale": parsed.get("importo_totale") or 0,
                "fornitore": {
                    "ragione_sociale": parsed.get("ragione_sociale_fornitore"),
                    "partita_iva": parsed.get("partita_iva_fornitore"),
                    "codice_fiscale": parsed.get("codice_fiscale_fornitore"),
                },
                "_fonte": "Parsing PDF (AI)",
            }
            return result

    except Exception as e:
        print(f"[Importa] Errore AI fattura: {e}")

    return None


# ═══════════════════════════════════════════════════════════════
# LETTURA DOCUMENTO PER CONTESTO AI (relazione, computo, ecc.)
# ═══════════════════════════════════════════════════════════════

def leggi_documento_per_contesto(file_path, max_chars=6000):
    """
    Legge un documento (PDF, DOCX, TXT) e ne estrae il testo
    da passare all'AI come contesto per generare le premesse.
    Ritorna il testo troncato a max_chars.
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".pdf":
        testo = estrai_testo_pdf(file_path, max_pagine=10)
    elif ext == ".docx":
        testo = _leggi_docx(file_path)
    elif ext in (".txt", ".md", ".csv"):
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            testo = f.read()
    elif ext in (".xlsx", ".xls"):
        testo = _leggi_excel_come_testo(file_path)
    else:
        raise ValueError(f"Formato non supportato: {ext}")

    # Tronca se troppo lungo
    if len(testo) > max_chars:
        testo = testo[:max_chars] + "\n\n[...troncato...]"

    return testo


def _leggi_docx(docx_path):
    """Estrae testo da un file .docx."""
    from docx import Document
    doc = Document(docx_path)
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _leggi_excel_come_testo(excel_path):
    """Converte un Excel in testo leggibile."""
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    lines = []
    for ws in wb.worksheets:
        lines.append(f"=== Foglio: {ws.title} ===")
        for row in ws.iter_rows(values_only=True):
            vals = [str(c) if c is not None else "" for c in row]
            if any(v.strip() for v in vals):
                lines.append(" | ".join(vals))
    wb.close()
    return "\n".join(lines)


def genera_premesse_da_documento(file_path, dati_affidamento):
    """
    Legge un documento (relazione, computo, ecc.) e usa l'AI
    per generare premesse contestualizzate.
    Ritorna il testo delle premesse, oppure None.
    """
    try:
        from ai_helper import get_api_key, MODEL
        import urllib.request
        import ssl

        api_key = get_api_key()
        if not api_key:
            return None

        testo_doc = leggi_documento_per_contesto(file_path)
        oggetto = dati_affidamento.get("oggetto", "")
        tipo = dati_affidamento.get("tipo_prestazione", "")

        prompt = f"""Sei un esperto di atti amministrativi del Comune di Sanremo, Servizio Demanio Marittimo.

Ti fornisco il testo di un documento tecnico (relazione, computo metrico, perizia) relativo a questo affidamento:
- Oggetto: {oggetto}
- Tipo prestazione: {tipo}

Basandoti sul documento, scrivi 3-4 premesse per la determina a contrarre.
Ogni premessa deve essere un periodo che termina con punto e virgola.
Scrivi in linguaggio amministrativo formale.
NON inventare dati non presenti nel documento.

DOCUMENTO:
{testo_doc}"""

        payload = json.dumps({
            "model": MODEL,
            "max_tokens": 2000,
            "messages": [{"role": "user", "content": prompt}],
            "system": "Sei un esperto di diritto amministrativo e atti della PA italiana. "
                      "Scrivi premesse formali per determinazioni dirigenziali.",
        }).encode("utf-8")

        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=payload,
            headers={
                "Content-Type": "application/json",
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01",
            },
            method="POST",
        )

        ctx = ssl.create_default_context()
        with urllib.request.urlopen(req, timeout=45, context=ctx) as resp:
            data = json.loads(resp.read().decode("utf-8"))

        if data.get("content"):
            return data["content"][0].get("text", "").strip()

    except Exception as e:
        print(f"[Importa] Errore AI documento: {e}")

    return None
